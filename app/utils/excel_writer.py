import io
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
import tempfile

def save_to_excel(results: list):
    """
    여러 장의 이미지 OCR 결과를 포함한 엑셀 파일을 생성.
    results: 리스트 형태의 dict들
        [
            {
                "filename": "image1.jpg",
                "image": <PIL.Image>,
                "text": "123가4567",
                "rerun_text": "123가4567"
            },
            ...
        ]
    """
    today_str = datetime.today().strftime("%Y-%m-%d")
    filename = f"차량번호_OCR결과_{today_str}.xlsx"

    # 1. 엑셀 데이터 구성
    data = []
    for idx, item in enumerate(results, start=1):
        data.append({
            "연번": idx,
            "사진": "이미지",  # 자리만 차지, 실제로는 이미지가 삽입됨
            "차량번호1": item["text"],
            "차량번호2": item["rerun_text"]
        })

    df = pd.DataFrame(data)

    # 2. 엑셀 저장 (BytesIO → openpyxl)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="OCR 결과")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # 3. 열 너비 조정 (B열: 약 47 = 350px 정도)
    col_widths = [8, 47, 20, 20]  # 연번 / 사진 / 차량번호1 / 차량번호2
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 4. 각 행에 이미지 삽입
    for idx, item in enumerate(results, start=2):  # 엑셀 데이터는 2행부터 시작
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            item["image"].save(tmp.name)
            img = XLImage(tmp.name)
            img.width = 350
            img.height = 170
            img.anchor = f'B{idx}'  # B열 + 해당 행
            ws.add_image(img)

            # 행 높이 조정 (엑셀 기준 약 170px → 127 row height 정도)
            ws.row_dimensions[idx].height = 127

    # 5. 최종 저장
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output, filename

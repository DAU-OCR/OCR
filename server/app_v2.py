import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ['CUDA_VISIBLE_DEVICES'] = ''
import json
import re
import io
import cv2
import torch
import easyocr
import numpy as np
import pandas as pd
import sys
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from PIL import Image as PILImage, ImageDraw, ImageFont, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import warnings
import traceback
from collections import Counter
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'yolov5'))
from utils.general import non_max_suppression, scale_boxes
from utils.augmentations import letterbox

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 환경 설정
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "1"
warnings.filterwarnings("ignore", category=FutureWarning)
cv2.setNumThreads(0)
cv2.ocl.setUseOpenCL(False)

# 디렉토리 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
CROP_FOLDER = os.path.join(UPLOAD_FOLDER, 'cropped')
VISUAL_FOLDER = os.path.join(UPLOAD_FOLDER, 'visual')
for d in [UPLOAD_FOLDER, CROP_FOLDER, VISUAL_FOLDER]:
    os.makedirs(d, exist_ok=True)
WARPED_FOLDER = os.path.join(UPLOAD_FOLDER, 'warped')
os.makedirs(WARPED_FOLDER, exist_ok=True)

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ✅ [추가] JSON 응답 시 한글 유니코드 이스케이프를 비활성화 (가장 중요)
app.config['JSON_AS_ASCII'] = False

@app.after_request
def cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

# 한글 매핑 딕셔너리
dict_map = {
    'gyeongnam': '경남', 'busan': '부산', 'jeo': '저', 'meo': '머', 'seo': '서', 'beo': '버',
    'geo': '거', 'neo': '너', 'deo': '더', 'bu': '부', 'do': '도', 'no': '노', 'go': '고', 'ro': '로',
    'bo': '보', 'jo': '조', 'gu': '구', 'na': '나', 'ma': '마', 'ba': '바', 'sa': '사', 'ah': '아',
    'ja': '자', 'cha': '차', 'ka': '카', 'ta': '타', 'pa': '파', 'ha': '하', 'la': '라', 'ra': '라',
    'me': '머', 'mu': '무', 'su': '수', 'ho': '호', 'ru': '루', 'mo': '모', 'ke': '커', 'ne': '네',
    'je': '제', 'yu': '유', 'se': '서', 'mi': '미', 'ju': '주', 'de': '데', 'oe': '외', 'wa': '와',
    'wi': '위', 'ri': '리', 'ye': '예', 'yi': '이', 'u': '우', 'eo': '어', 'heo': '허', 'du': '두',
    'leo': '러', 'lu': '루', 'so': '소', 'da': '다', 'lo': '로', 'nu': '누', 'o': '오', 'ga': '가'
}
dict_sorted = sorted(dict_map.items(), key=lambda x: len(x[0]), reverse=True)

# 번호판에 사용되는 모든 한글 문자 집합
VALID_HANGUL_CHARS = {
    # Private vehicles
    '가', '나', '다', '라', '마', '거', '너', '더', '러', '머', '버', '서', '어', '저',
    '고', '노', '도', '로', '모', '보', '소', '오', '조', '구', '누', '두', '루', '무',
    '부', '수', '우', '주',
    # Rental cars
    '허', '하', '호',
    # Commercial vehicles (Taxis, Buses)
    '바', '사', '아', '자',
    # Delivery
    '배',
    # Military (optional)
    '육', '해', '공', '국', '합'
}

# OCR 유틸 함수
def roman_to_korean(text):
    t = text.lower()
    for roman, kor in dict_sorted:
        t = t.replace(roman, kor)
    return t

def normalize(text):
    return re.sub(r'[^가-힣0-9]', '', text)

def is_valid_plate(text):
    return bool(re.fullmatch(r'\d{2,3}[가-힣]\d{4}', text))

def reorder_blocks(blocks):
    if len(blocks) == 2:
        if re.search(r'[가-힣]', blocks[1]) and not re.search(r'[가-힣]', blocks[0]):
            return [blocks[1], blocks[0]]
    return blocks

def insert_hangul_fixed(digits: str, hangul: str) -> str:
    if len(digits) >= 7:
        pos = -5
        return digits[:pos] + hangul + digits[pos+1:]
    return digits

def patch_hangul(t1, t2):
    d1 = ''.join(re.findall(r'\d', t1))
    d2 = ''.join(re.findall(r'\d', t2))
    h1 = re.findall(r'[가-힣]', t1)
    h2 = re.findall(r'[가-힣]', t2)
    if len(d1) in [7, 8] and len(h2) == 1:
        p = insert_hangul_fixed(d1, h2[0])
        if is_valid_plate(p): return p
    if len(d2) in [7, 8] and len(h1) == 1:
        p = insert_hangul_fixed(d2, h1[0])
        if is_valid_plate(p): return p
    return None

def get_filtered_ocr(reader, image, resize):
    resized = cv2.resize(image, resize)
    result = reader.readtext(resized)
    if not result:
        return '', 0.0
    blocks = reorder_blocks([t for (_, t, _) in result])
    merged = ''.join(blocks)
    norm = normalize(merged)
    conf = max([c for (_, _, c) in result])
    return norm, round(conf, 2)

def apply_plate_selection_logic(t1, c1, t2, c2, t3, c3, hangul_dict):
    results = [
        {'text': t1, 'conf': c1, 'name': '모델1'},
        {'text': t2, 'conf': c2, 'name': '모델2'},
        {'text': t3, 'conf': c3, 'name': '모델3(CRNN)'}
    ]

    # --- 1단계: 다수결 투표 ---
    # 정규식과 한글 사전을 모두 통과한 결과만 투표 자격 부여
    vote_candidates = []
    for r in results:
        if is_valid_plate(r['text']):
            hangul_char = re.findall(r'[가-힣]', r['text'])
            if hangul_char and hangul_char[0] in hangul_dict:
                vote_candidates.append(r['text'])
    
    if len(vote_candidates) >= 2:
        vote_counts = Counter(vote_candidates)
        most_common = vote_counts.most_common(1)[0]
        if most_common[1] >= 2: # 2개 이상 동의
            return most_common[0], f"다수결({most_common[1]}표)"

    # --- 2단계: 모델1 우선 규칙 ---
    is_t1_valid = is_valid_plate(t1)
    if is_t1_valid:
        hangul_char = re.findall(r'[가-힣]', t1)
        if hangul_char and hangul_char[0] in hangul_dict:
            return t1, "모델1우선"

    # --- 3단계: 가중치 적용 비교 ---
    # 2단계에서 모델1이 채택되지 않았으므로, 모델1을 제외한 나머지로 비교
    remaining_results = [r for r in results if r['name'] != '모델1']
    valid_results = [r for r in remaining_results if is_valid_plate(r['text'])]
    
    if valid_results:
        for r in valid_results:
            hangul_char = re.findall(r'[가-힣]', r['text'])
            r['in_dict'] = hangul_char and hangul_char[0] in hangul_dict

        dict_results = [r for r in valid_results if r['in_dict']]

        def get_adjusted_conf(r):
            if r['name'] == '모델3(CRNN)':
                return r['conf'] * 0.9
            return r['conf']

        if dict_results:
            best = max(dict_results, key=get_adjusted_conf)
            return best['text'], f"사전+가중치({best['name']})"
        else:
            best = max(valid_results, key=get_adjusted_conf)
            return best['text'], f"정규식+가중치({best['name']})"

    # --- 4단계: 패치 및 최종 선택 ---
    patch_pairs = [(t1, t2), (t1, t3), (t2, t3)]
    for p1, p2 in patch_pairs:
        patched = patch_hangul(p1, p2)
        if patched and is_valid_plate(patched):
            hangul_char = re.findall(r'[가-힣]', patched)
            if hangul_char and hangul_char[0] in hangul_dict:
                return patched, '패치'

    def get_adjusted_conf_final(r):
        if r['name'] == '모델3(CRNN)':
            return r['conf'] * 0.9
        return r['conf']
    best = max(results, key=get_adjusted_conf_final)
    return best['text'], f"conf+가중치({best['name']})"

def get_plate_corners(image, fname=None, save_debug=False, debug_dir=None):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blur, 50, 150)

    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'gray.png'), gray)
        cv2.imwrite(os.path.join(debug_dir, 'blur.png'), blur)
        cv2.imwrite(os.path.join(debug_dir, 'canny.png'), edged)

    contours, _ = cv2.findContours(edged, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    contour_vis = image.copy()
    cv2.drawContours(contour_vis, contours, -1, (0, 255, 0), 1)
    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'contours.png'), contour_vis)

    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    for cnt in contours[:10]:
        area = cv2.contourArea(cnt)
        if area < 1000:
            continue
        approx = cv2.approxPolyDP(cnt, 0.02 * cv2.arcLength(cnt, True), True)
        if len(approx) == 4:
            pts = approx.reshape(4, 2)
            x, y, w, h = cv2.boundingRect(pts)
            if w / h < 2:
                continue
            return order_corners(pts)
    return None


def get_plate_corners_threshold(image, fname=None, save_debug=False, debug_dir=None):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 180, 255, cv2.THRESH_BINARY)

    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'blur_thresh.png'), blur)
        cv2.imwrite(os.path.join(debug_dir, 'thresh.png'), thresh)

    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contour_vis = image.copy()
    cv2.drawContours(contour_vis, contours, -1, (0, 0, 255), 1)
    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'contours_thresh.png'), contour_vis)

    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < 1500:
            continue
        rect = cv2.minAreaRect(cnt)
        box = cv2.boxPoints(rect)
        ordered = order_corners(box.astype(np.float32))
        if save_debug and debug_dir:
            box_img = image.copy()
            cv2.polylines(box_img, [np.int32(ordered)], True, (255, 0, 0), 2)
            cv2.imwrite(os.path.join(debug_dir, 'minAreaRect_box.png'), box_img)
        return ordered
    return None



def order_corners(pts):
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)] # top-left
    rect[2] = pts[np.argmax(s)] # bottom-right

    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)] # top-right
    rect[3] = pts[np.argmax(diff)] # bottom-left
    return rect

def warp_perspective(image, corners, output_size=(200, 60)):
    dst = np.array([
        [0, 0],
        [output_size[0]-1, 0],
        [output_size[0]-1, output_size[1]-1],
        [0, output_size[1]-1]
    ], dtype="float32")
    M = cv2.getPerspectiveTransform(corners, dst)
    return cv2.warpPerspective(image, M, output_size)


# 모델 로딩
yolo_model = torch.hub.load(os.path.join(BASE_DIR, 'yolov5'), 'custom',
                             path=resource_path(os.path.join('custom_weights', 'best.pt')),
                             source='local',
                             device='cpu', # 이 부분을 추가합니다.
                             verbose=False)

model_path = resource_path('custom_weights_easyOCR')
reader1 = easyocr.Reader(['ko'], gpu=False,
                         model_storage_directory=model_path,
                         user_network_directory=model_path,
                         recog_network='korean_g2', download_enabled=False)
reader2 = easyocr.Reader(['en'], gpu=False,
                         model_storage_directory=model_path,
                         user_network_directory=model_path,
                         recog_network='best_acc', download_enabled=False)

# --- CRNN 모델 로딩 ---
device = torch.device('cpu')
crnn_model_path = resource_path(os.path.join('CRNN_model', 'ocrBestModel_142.pth'))

# CRNN_model 폴더를 시스템 경로에 추가
sys.path.insert(0, resource_path('CRNN_model'))

from preprocess import load_crnn_model, run_crnn_ocr
from label_encoder import LabelEncoder

label_encoder = LabelEncoder()
charset = label_encoder.get_charset()
num_classes = len(charset) + 1 # CTC blank 토큰을 위해 +1

crnn_model = load_crnn_model(crnn_model_path, num_classes=num_classes, device=device)
# --- CRNN 모델 로딩 종료 ---

MIN_AREA = 1400
MIN_ASPECT_RATIO = 1.0
resize1 = (100, 32)
resize2 = (200, 60)
records = []
def detect_plate(image):
    # Preprocessing
    img0 = image.copy()
    img, ratio, (dw, dh) = letterbox(img0, 640, stride=32, auto=True)
    img = img.transpose((2, 0, 1))[::-1] # HWC to CHW, BGR to RGB
    img = np.ascontiguousarray(img)

    img = torch.from_numpy(img).to(device)
    img = img.float()
    img /= 255.0
    if len(img.shape) == 3:
        img = img[None]

    # Inference
    # pred = yolo_model(img, augment=False, visualize=False)    수정
    pred = yolo_model(img, augment=False ) # visualize 옵션 제거
    if isinstance(pred, (list, tuple)) and len(pred) in (2, 3): # yolov5 returns a tuple (pred, feature_maps) or (pred, protos, feature_maps)
        pred = pred[0]

    # NMS
    pred = non_max_suppression(pred, 0.25, 0.45, classes=None, agnostic=False, max_det=1000)

    detections = []
    for i, det in enumerate(pred): # per image
        if len(det):
            # Rescale boxes from img_size to im0 size
            det[:, :4] = scale_boxes(img.shape[2:], det[:, :4], img0.shape).round()
            detections = det.cpu().numpy()

    if len(detections) == 0:
        return None

    filtered = []
    for *xyxy, conf, cls in detections.tolist():
        x1, y1, x2, y2 = map(int, xyxy)
        area = (x2 - x1) * (y2 - y1)
        ratio_bbox = (x2 - x1) / (y2 - y1 + 1e-5)
        if area > MIN_AREA and ratio_bbox > MIN_ASPECT_RATIO:
            filtered.append((area, (x1, y1, x2, y2)))

    if not filtered:
        return None

    _, (x1, y1, x2, y2) = max(filtered)
    return image[y1:y2, x1:x2], (x1, y1, x2, y2)

def process_file_and_record(f, fname, image):
    """
    OCR 처리 및 records에 결과 저장 로직을 담은 헬퍼 함수
    /upload 와 /upload-batch 에서 재사용
    """
    try:
        result = {
            'image': f'/uploads/{fname}',
            'timestamp': datetime.now().isoformat(),
            'matched': False
        }

        # 디버그 폴더 생성
        debug_subdir = os.path.join(UPLOAD_FOLDER, 'debug', os.path.splitext(fname)[0])
        os.makedirs(debug_subdir, exist_ok=True)

        detected = detect_plate(image)
        if not detected:
            records.append(result)
            return

        plate_img, (x1, y1, x2, y2) = detected

        # 디버그: YOLO 검출 영역 저장
        cv2.imwrite(os.path.join(debug_subdir, '1_detected_yolo.png'), plate_img)

        # 패딩 추가
        pad = 20
        x1 = max(x1 - pad, 0)
        y1 = max(y1 - pad, 0)
        x2 = min(x2 + pad, image.shape[1])
        y2 = min(y2 + pad, image.shape[0])
        plate_img = image[y1:y2, x1:x2]

        # 디버그: 패딩 후 이미지 저장
        cv2.imwrite(os.path.join(debug_subdir, '2_padded.png'), plate_img)

        # 디버그: 보정 전 이미지 저장
        cv2.imwrite(os.path.join(debug_subdir, '3_before_correction.png'), plate_img)

        # 보정 시도 1
        corners = get_plate_corners(plate_img, fname=os.path.splitext(fname)[0], save_debug=True, debug_dir=debug_subdir)
        if corners is None:
            corners = get_plate_corners_threshold(plate_img, fname=os.path.splitext(fname)[0], save_debug=True, debug_dir=debug_subdir)


        if corners is not None:
            # 디버그: 검출된 코너 시각화
            corner_img = plate_img.copy()
            cv2.polylines(corner_img, [np.int32(corners)], True, (0, 255, 0), 2)
            cv2.imwrite(os.path.join(debug_subdir, '4_detected_corners.png'), corner_img)

            # 보정 수행
            warped_img = warp_perspective(plate_img, corners)

            # 디버그: 보정된 결과 이미지 저장
            cv2.imwrite(os.path.join(debug_subdir, '5_corrected_warped.png'), warped_img)

            # 보정된 이미지로 교체
            plate_img = warped_img

        # OCR 리사이즈 후 입력 저장
        t1_input = cv2.resize(plate_img, resize1)
        t2_input = cv2.resize(plate_img, resize2)
        cv2.imwrite(os.path.join(debug_subdir, '6_ocr_model1_input.png'), t1_input)
        cv2.imwrite(os.path.join(debug_subdir, '7_ocr_model2_input.png'), t2_input)

        # OCR 수행
        t1, c1 = get_filtered_ocr(reader1, plate_img, resize1)
        t2, c2 = get_filtered_ocr(reader2, plate_img, resize2)
        t3, c3 = run_crnn_ocr(plate_img, crnn_model, label_encoder, device)
        c3 = round(c3, 2)

        selected, reason = apply_plate_selection_logic(t1, c1, t2, c2, t3, c3, VALID_HANGUL_CHARS)
        matched = is_valid_plate(selected)

        if not matched:
            selected = '인식 실패'

        crop_name = f"crop_{fname}"
        vis_name = f"vis_{fname}"
        cv2.imwrite(os.path.join(CROP_FOLDER, crop_name), plate_img)
        PILImage.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB)).save(os.path.join(VISUAL_FOLDER, vis_name))

        # 개발자용 디버깅 이미지 저장: YOLO 박스 포함된 원본 이미지
        DEV_VISUAL_FOLDER = os.path.join(UPLOAD_FOLDER, 'dev_visual')
        os.makedirs(DEV_VISUAL_FOLDER, exist_ok=True)
        dev_vis_path = os.path.join(DEV_VISUAL_FOLDER, f'dev_{fname}')
        dev_img = image.copy()
        cv2.rectangle(dev_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        PILImage.fromarray(cv2.cvtColor(dev_img, cv2.COLOR_BGR2RGB)).save(dev_vis_path)

        result.update({
            'crop': f'/uploads/cropped/{crop_name}',
            'visual': f'/uploads/visual/{vis_name}',
            'text1': t1, 'conf1': c1,
            'text2': t2, 'conf2': c2,
            'text3': t3, 'conf3': c3,
            'plate': selected, 'reason': reason,
            'matched': matched
        })
        records.append(result)

    except Exception as e:
        print(f"Error processing file {fname}: {e}")
        traceback.print_exc()
        # 오류 발생 시 실패 결과만 기록하고 계속 진행
        records.append({
            'image': f'/uploads/{fname}',
            'timestamp': datetime.now().isoformat(),
            'matched': False,
            'plate': '처리 오류',
            'reason': f'Internal Error: {str(e)}'
        })


# =========================================================================
# 1. 기존 웹 UI 업로드 엔드포인트 유지
# =========================================================================
@app.route('/upload', methods=['POST'])
def upload():
    try:
        # 기존 웹 UI와 호환되는 'images' 또는 'image' 필드를 사용
        files = request.files.getlist('images') or request.files.getlist('image')
        if not files:
            return jsonify({'error': '파일이 없습니다'}), 400

        for f in files:
            fname = secure_filename(f.filename)
            path = os.path.join(UPLOAD_FOLDER, fname)
            f.save(path)
            image = cv2.imread(path)
            
            # OCR 및 기록 로직 재사용
            process_file_and_record(f, fname, image)

        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        # Return a 500 error
        return jsonify({'error': 'An internal error occurred', 'details': str(e)}), 500


# =========================================================================
# 2. BAT 파일용 일괄 업로드 엔드포인트 추가 (새로운 기능)
# =========================================================================
@app.route('/upload-batch', methods=['POST'])
def upload_batch():
    try:
        # BAT 파일의 curl 명령 (-F "files=@...")과 일치하도록 'files' 필드를 사용
        files = request.files.getlist('files')
        if not files:
            # BAT 파일의 오류 처리를 위해 400 Bad Request를 반환
            return jsonify({'error': '업로드할 이미지 파일이 없습니다.'}), 400

        for f in files:
            fname = secure_filename(f.filename)
            path = os.path.join(UPLOAD_FOLDER, fname)
            f.save(path)
            image = cv2.imread(path)
            
            # OCR 및 기록 로직 재사용
            process_file_and_record(f, fname, image)
            
        # BAT 파일이 다음 단계(다운로드)로 진행할 수 있도록 성공 응답 반환
        return jsonify({'status': 'batch upload ok', 'count': len(files)}), 200

    except Exception as e:
        print(f"Batch Upload Error: {e}")
        traceback.print_exc()
        # 오류 발생 시 500 에러를 반환하여 BAT 파일이 인식하도록 함
        return jsonify({'error': 'An internal error occurred during batch processing', 'details': str(e)}), 500


# =========================================================================
# 3. 나머지 엔드포인트는 그대로 유지
# =========================================================================
@app.route('/download', methods=['GET'])
def download_excel():
    # ... (기존 코드 유지) ...
    data = records.copy()

    if not data:
        return jsonify({'error': '데이터 없음'}), 400

    excel_data = []
    for r in data:
        # 모델 결과를 신뢰도 순으로 정렬하여 엑셀에 넣을 데이터를 준비
        all_results = [
            {'name': '모델1', 'text': r.get('text1', ''), 'conf': r.get('conf1', 0)},
            {'name': '모델2', 'text': r.get('text2', ''), 'conf': r.get('conf2', 0)},
            {'name': '모델3(CRNN)', 'text': r.get('text3', ''), 'conf': r.get('conf3', 0)}
        ]
        sorted_results = sorted(all_results, key=lambda x: x['conf'], reverse=True)
        
        # 실패했을 경우 텍스트를 '인식 실패'로 통일
        plate_text = r.get('plate', 'N/A')
        if not r.get('matched') or plate_text == '인식 실패':
            top1_text = '인식 실패'
            top2_text = '인식 실패'
            plate_text = '인식 실패'
        else:
            top1_text = sorted_results[0]['text']
            top2_text = sorted_results[1]['text']
        
        row_data = {
            'Top1_Text': top1_text,
            'Top2_Text': top2_text,
            'plate': plate_text,
            'image_path': r['image'] # 이미지 경로 추가 (엑셀 삽입용)
        }
        excel_data.append(row_data)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # DataFrame 생성 시 이미지 경로를 제외한 컬럼만 사용
        df = pd.DataFrame(excel_data)[['Top1_Text', 'Top2_Text', 'plate']]
        df.columns = ['1순위 모델 결과', '2순위 모델 결과', '선택된 결과']
        df.to_excel(writer, index=False, sheet_name='결과')
        ws = writer.sheets['결과']

        # 이미지 열 추가
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = '차량 이미지'

        TARGET_WIDTH = 450 # 가로 너비 기준을 450px로 설정 (원하는 크기로 조정)

        def px_to_col_width(px): return px * 0.14
        def px_to_row_height(px): return px * 0.75

        # 엑셀 데이터 대신 원본 records를 사용하여 이미지 경로에 접근
        for idx, r in enumerate(data, start=2):
            try:
                # 이미지 경로는 원본 'image' 필드를 사용
                img_path = os.path.join(BASE_DIR, r['image'].lstrip('/'))
                pil = PILImage.open(img_path)

                # 이미지 회전 문제 해결 로직 적용
                # EXIF 메타데이터 기반으로 이미지 자동 회전 보정
                pil = ImageOps.exif_transpose(pil)
                
                orig_width, orig_height = pil.size

                # 비율 유지한 세로 크기 계산
                scale = TARGET_WIDTH / orig_width
                resized_height = int(orig_height * scale)

                # 이미지 리사이즈 (PIL.LANCZOS는 고품질 리사이즈 필터)
                resized_img = pil.resize((TARGET_WIDTH, resized_height), PILImage.LANCZOS)
                
                img_bytes = io.BytesIO()
                resized_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)

                xl_img = XLImage(img_bytes)
                xl_img.width = TARGET_WIDTH
                xl_img.height = resized_height

                ws.add_image(xl_img, f'A{idx}')
                ws.row_dimensions[idx].height = px_to_row_height(resized_height)

            except Exception as e:
                # 오류 발생 시 이미지 셀 건너뛰기
                print(f"[이미지 삽입 실패] {r.get('image', '')} → {e}")

        ws.column_dimensions['A'].width = px_to_col_width(TARGET_WIDTH)

        # 컬럼 너비 조정 (총 4개 컬럼: 이미지, 1순위, 2순위, 최종)
        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    buf.seek(0)
    today = datetime.now().strftime('%Y-%m-%d')
    fname = secure_filename(request.args.get('filename') or f"{today}_plates") + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/results')
def results():
    return jsonify(records), 200

@app.route('/update-plates', methods=['POST'])
def update_plates():
    updates = request.json # 리스트 받기
    count = 0
    for u in updates:
        for r in records:
            if r['image'] == u['image']:
                r['plate'] = u['plate']
                r['matched'] = is_valid_plate(u['plate'])
                count += 1
    return jsonify({'updated': count}), 200



@app.route('/reset', methods=['POST'])
def reset():
    records.clear()
    return jsonify({'status': 'ok'})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route('/download-json', methods=['GET'])
def download_json():
    data = records.copy()
    if not data:
        # 데이터가 없을 경우 jsonify를 사용 (여기서는 한글이 포함된 데이터가 아니므로 무방함)
        return jsonify({'error': '데이터 없음'}), 400

    output_data = []
    for r in data:
        selected_conf = 0
        accuracy = "N/A"
        
        reason_text = r.get('reason', '')
        model_match = re.search(r'\((.*?)\)', reason_text)
        selected_model_name = model_match.group(1) if model_match else ''

        if selected_model_name == '모델1':
            selected_conf = r.get('conf1', 0)
        elif selected_model_name == '모델2':
            selected_conf = r.get('conf2', 0)
        elif selected_model_name == '모델3(CRNN)':
            selected_conf = r.get('conf3', 0)

        if '패치' in reason_text:
            accuracy = "N/A (Patched)"
        elif selected_conf > 0:
            accuracy = f"{selected_conf * 100:.2f}"
        
        error_message = ""
        if not r.get('matched') or r.get('plate') == '인식 실패':
            error_message = reason_text or '인식 실패'

        json_record = {
            '파일명': os.path.basename(r.get('image', '')),
            '처리일시': r.get('timestamp'),
            '모델별 결과': [
                {'모델명': '모델1 (EasyOCR-ko)', '결과': r.get('text1', ''), '신뢰도': r.get('conf1', 0)},
                {'모델명': '모델2 (EasyOCR-en)', '결과': r.get('text2', ''), '신뢰도': r.get('conf2', 0)},
                {'모델명': '모델3 (CRNN)', '결과': r.get('text3', ''), '신뢰도': r.get('conf3', 0)},
            ],
            '최종선택결과': r.get('plate'),
            '정확도(%)': accuracy,
            '오류메시지': error_message
        }
        output_data.append(json_record)

    # -----------------------------------------------------------
    # ✅ [수정] JSON 직렬화 강제 적용
    # JSON_AS_ASCII=False 설정이 작동하지 않을 경우를 대비하여
    # python의 json.dumps를 사용하여 ensure_ascii=False를 명시합니다.
    # -----------------------------------------------------------
    json_string = json.dumps(output_data, ensure_ascii=False, indent=4)
    
    # Flask Response 객체를 직접 생성하여 Content-Type과 인코딩을 명확히 지정
    response = app.response_class(
        response=json_string,
        status=200,
        mimetype='application/json; charset=utf-8'
    )
    
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"ocr_results_{today}.json"
    
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
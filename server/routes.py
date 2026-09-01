import io
import json
import os
import re
import traceback
import zipfile
from datetime import datetime

import cv2
import numpy as np
import pandas as pd
from flask import Blueprint, request, jsonify, send_file, send_from_directory
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps
from werkzeug.utils import secure_filename

from config import BASE_DIR, UPLOAD_FOLDER
from ocr_pipeline import process_file_and_record
from state import records
from text_utils import is_valid_plate

bp = Blueprint('api', __name__)


# =========================================================================
# 1. 기존 웹 UI 업로드 엔드포인트 유지
# =========================================================================
@bp.route('/upload', methods=['POST'])
def upload():
    try:
        # 기존 웹 UI와 호환되는 'images' 또는 'image' 필드를 사용
        files = request.files.getlist('images') or request.files.getlist('image')
        if not files:
            return jsonify({'error': '파일이 없습니다'}), 400

        total_processed_from_zip = 0
        total_images_in_zip = 0
        for f in files:
            fname = secure_filename(f.filename)
            if fname.lower().endswith('.zip'):
                zip_buffer = io.BytesIO(f.read())
                with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
                    # Count actual image files in zip for total_images_in_zip
                    zip_image_files = [zi for zi in zip_ref.infolist() if not zi.is_dir() and zi.filename.lower().endswith(('.png', '.jpg', '.jpeg'))]
                    total_images_in_zip = len(zip_image_files)

                    for zip_info in zip_image_files:
                        image_data = zip_ref.read(zip_info.filename)
                        image_np = np.frombuffer(image_data, np.uint8)
                        image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)

                        unique_zip_fname = secure_filename(zip_info.filename.replace(os.sep, '_').replace('/', '_'))

                        path = os.path.join(UPLOAD_FOLDER, unique_zip_fname)
                        with open(path, 'wb') as img_file:
                            img_file.write(image_data)

                        process_file_and_record(None, unique_zip_fname, image)
                        total_processed_from_zip += 1
                return jsonify({'status': 'ok', 'processed_count': total_processed_from_zip, 'total_images_in_zip': total_images_in_zip}), 200
            else:
                path = os.path.join(UPLOAD_FOLDER, fname)
                f.seek(0)
                f.save(path)
                image = cv2.imread(path)

                process_file_and_record(f, fname, image)
                return jsonify({'status': 'ok', 'processed_count': 1}), 200

        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        # Return a 500 error
        return jsonify({'error': 'An internal error occurred', 'details': str(e)}), 500


# =========================================================================
# 2. BAT 파일용 일괄 업로드 엔드포인트 추가 (새로운 기능)
# =========================================================================
@bp.route('/upload-batch', methods=['POST'])
def upload_batch():
    try:
        # BAT 파일의 curl 명령 (-F "files=@...")과 일치하도록 'files' 필드를 사용
        files = request.files.getlist('files')
        if not files:
            # BAT 파일의 오류 처리를 위해 400 Bad Request를 반환
            return jsonify({'error': '업로드할 이미지 파일이 없습니다.'}), 400

        processed_count = 0
        for f in files:
            fname = secure_filename(f.filename)
            if fname.lower().endswith('.zip'):
                zip_buffer = io.BytesIO(f.read())
                with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
                    for zip_info in zip_ref.infolist():
                        if zip_info.is_dir() or not zip_info.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                            continue

                        image_data = zip_ref.read(zip_info.filename)
                        image_np = np.frombuffer(image_data, np.uint8)
                        image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)

                        # Ensure unique filename by replacing path separators with underscores
                        unique_zip_fname = secure_filename(zip_info.filename.replace(os.sep, '_').replace('/', '_'))

                        path = os.path.join(UPLOAD_FOLDER, unique_zip_fname)
                        with open(path, 'wb') as img_file:
                            img_file.write(image_data)

                        process_file_and_record(None, unique_zip_fname, image)
                        processed_count += 1
            else:
                path = os.path.join(UPLOAD_FOLDER, fname)
                f.seek(0)
                f.save(path)
                image = cv2.imread(path)

                # OCR 및 기록 로직 재사용
                process_file_and_record(f, fname, image)
                processed_count += 1

        # BAT 파일이 다음 단계(다운로드)로 진행할 수 있도록 성공 응답 반환
        return jsonify({'status': 'batch upload ok', 'count': processed_count}), 200

    except Exception as e:
        print(f"Batch Upload Error: {e}")
        traceback.print_exc()
        # 오류 발생 시 500 에러를 반환하여 BAT 파일이 인식하도록 함
        return jsonify({'error': 'An internal error occurred during batch processing', 'details': str(e)}), 500


# =========================================================================
# 3. 나머지 엔드포인트는 그대로 유지
# =========================================================================
@bp.route('/download', methods=['GET'])
def download_excel():
    # ... (기존 코드 유지) ...
    data = records.copy()

    if not data:
        return jsonify({'error': '데이터 없음'}), 400

    excel_data = []
    for r in data:
        # 모델 결과를 신뢰도 순으로 정렬하여 엑셀에 넣을 데이터를 준비
        all_results = [
            {'name': '모델1', 'text': r.get('text1', ''), 'conf': r.get('conf1', 0)},
            {'name': '모델2', 'text': r.get('text2', ''), 'conf': r.get('conf2', 0)},
            {'name': '모델3(CRNN)', 'text': r.get('text3', ''), 'conf': r.get('conf3', 0)}
        ]
        sorted_results = sorted(all_results, key=lambda x: x['conf'], reverse=True)

        # 실패했을 경우 텍스트를 '인식 실패'로 통일
        plate_text = r.get('plate', 'N/A')
        if not r.get('matched') or plate_text == '인식 실패':
            top1_text = '인식 실패'
            top2_text = '인식 실패'
            plate_text = '인식 실패'
        else:
            top1_text = sorted_results[0]['text']
            top2_text = sorted_results[1]['text']

        row_data = {
            'Top1_Text': top1_text,
            'Top2_Text': top2_text,
            'plate': plate_text,
            'image_path': r['image'] # 이미지 경로 추가 (엑셀 삽입용)
        }
        excel_data.append(row_data)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # DataFrame 생성 시 이미지 경로를 제외한 컬럼만 사용
        df = pd.DataFrame(excel_data)[['Top1_Text', 'Top2_Text', 'plate']]
        df.columns = ['1순위 모델 결과', '2순위 모델 결과', '선택된 결과']
        df.to_excel(writer, index=False, sheet_name='결과')
        ws = writer.sheets['결과']

        # 이미지 열 추가
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = '차량 이미지'

        TARGET_WIDTH = 450 # 가로 너비 기준을 450px로 설정 (원하는 크기로 조정)

        def px_to_col_width(px): return px * 0.14
        def px_to_row_height(px): return px * 0.75

        # 엑셀 데이터 대신 원본 records를 사용하여 이미지 경로에 접근
        for idx, r in enumerate(data, start=2):
            try:
                # 이미지 경로는 원본 'image' 필드를 사용
                img_path = os.path.join(BASE_DIR, r['image'].lstrip('/'))
                pil = PILImage.open(img_path)

                # 이미지 회전 문제 해결 로직 적용
                # EXIF 메타데이터 기반으로 이미지 자동 회전 보정
                pil = ImageOps.exif_transpose(pil)

                orig_width, orig_height = pil.size

                # 비율 유지한 세로 크기 계산
                scale = TARGET_WIDTH / orig_width
                resized_height = int(orig_height * scale)

                # 이미지 리사이즈 (PIL.LANCZOS는 고품질 리사이즈 필터)
                resized_img = pil.resize((TARGET_WIDTH, resized_height), PILImage.LANCZOS)

                img_bytes = io.BytesIO()
                resized_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)

                xl_img = XLImage(img_bytes)
                xl_img.width = TARGET_WIDTH
                xl_img.height = resized_height

                ws.add_image(xl_img, f'A{idx}')
                ws.row_dimensions[idx].height = px_to_row_height(resized_height)

            except Exception as e:
                # 오류 발생 시 이미지 셀 건너뛰기
                print(f"[이미지 삽입 실패] {r.get('image', '')} → {e}")

        ws.column_dimensions['A'].width = px_to_col_width(TARGET_WIDTH)

        # 컬럼 너비 조정 (총 4개 컬럼: 이미지, 1순위, 2순위, 최종)
        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    buf.seek(0)
    today = datetime.now().strftime('%Y-%m-%d')
    fname = secure_filename(request.args.get('filename') or f"{today}_plates") + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/results')
def results():
    return jsonify(records), 200

@bp.route('/update-plates', methods=['POST'])
def update_plates():
    updates = request.json # 리스트 받기
    count = 0
    for u in updates:
        for r in records:
            if r['image'] == u['image']:
                r['plate'] = u['plate']
                r['matched'] = is_valid_plate(u['plate'])
                count += 1
    return jsonify({'updated': count}), 200


@bp.route('/reset', methods=['POST'])
def reset():
    records.clear()
    return jsonify({'status': 'ok'})

@bp.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


@bp.route('/download-json', methods=['GET'])
def download_json():
    data = records.copy()
    if not data:
        return jsonify({'error': '데이터 없음'}), 400

    output_data = []
    for r in data:
        selected_conf = 0
        accuracy = "N/A"

        reason_text = r.get('reason', '')
        model_match = re.search(r'\((.*?)\)', reason_text)
        selected_model_name = model_match.group(1) if model_match else ''

        if selected_model_name == '모델1':
            selected_conf = r.get('conf1', 0)
        elif selected_model_name == '모델2':
            selected_conf = r.get('conf2', 0)
        elif selected_model_name == '모델3(CRNN)':
            selected_conf = r.get('conf3', 0)

        if '패치' in reason_text:
            accuracy = "N/A (Patched)"
        elif selected_conf > 0:
            accuracy = f"{selected_conf * 100:.2f}"

        error_message = ""
        if not r.get('matched') or r.get('plate') == '인식 실패':
            error_message = reason_text or '인식 실패'

        json_record = {
            '파일명': os.path.basename(r.get('image', '')),
            '처리일시': r.get('timestamp'),
            '모델별 결과': [
                {'모델명': '모델1 (EasyOCR-ko)', '결과': r.get('text1', ''), '신뢰도': r.get('conf1', 0)},
                {'모델명': '모델2 (EasyOCR-en)', '결과': r.get('text2', ''), '신뢰도': r.get('conf2', 0)},
                {'모델명': '모델3 (CRNN)', '결과': r.get('text3', ''), '신뢰도': r.get('conf3', 0)},
            ],
            '최종선택결과': r.get('plate'),
            '정확도(%)': accuracy,
            '오류메시지': error_message
        }
        output_data.append(json_record)

    json_string = json.dumps(output_data, ensure_ascii=False, indent=4)

    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"ocr_results_{today}.json"

    response = send_file(
        io.BytesIO(json_string.encode('utf-8')),
        as_attachment=True,
        download_name=filename,
        mimetype='application/json'
    )
    print(f"[DEBUG] Content-Disposition header: {response.headers.get('Content-Disposition')}")
    return response

@bp.route('/get-zip-image-count', methods=['POST'])
def get_zip_image_count():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and file.filename.lower().endswith('.zip'):
        zip_buffer = io.BytesIO(file.read())
        try:
            with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
                image_count = 0
                for zip_info in zip_ref.infolist():
                    if not zip_info.is_dir() and zip_info.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                        image_count += 1
                return jsonify({'image_count': image_count}), 200
        except zipfile.BadZipFile:
            return jsonify({'error': 'Bad zip file'}), 400
    return jsonify({'error': 'Not a zip file'}), 400

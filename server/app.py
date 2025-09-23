import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import re
import io
import cv2
import torch
import easyocr
import numpy as np
import pandas as pd
import sys
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import warnings

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 환경 설정
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "1"
os.environ['CUDA_VISIBLE_DEVICES'] = ''
warnings.filterwarnings("ignore", category=FutureWarning)
cv2.setNumThreads(0)
cv2.ocl.setUseOpenCL(False)

# 디렉토리 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
CROP_FOLDER = os.path.join(UPLOAD_FOLDER, 'cropped')
VISUAL_FOLDER = os.path.join(UPLOAD_FOLDER, 'visual')
for d in [UPLOAD_FOLDER, CROP_FOLDER, VISUAL_FOLDER]:
    os.makedirs(d, exist_ok=True)
WARPED_FOLDER = os.path.join(UPLOAD_FOLDER, 'warped')
os.makedirs(WARPED_FOLDER, exist_ok=True)

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.after_request
def cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

# 한글 매핑 딕셔너리
dict_map = {
    'gyeongnam': '경남', 'busan': '부산', 'jeo': '저', 'meo': '머', 'seo': '서', 'beo': '버',
    'geo': '거', 'neo': '너', 'deo': '더', 'bu': '부', 'do': '도', 'no': '노', 'go': '고', 'ro': '로',
    'bo': '보', 'jo': '조', 'gu': '구', 'na': '나', 'ma': '마', 'ba': '바', 'sa': '사', 'ah': '아',
    'ja': '자', 'cha': '차', 'ka': '카', 'ta': '타', 'pa': '파', 'ha': '하', 'la': '라', 'ra': '라',
    'me': '머', 'mu': '무', 'su': '수', 'ho': '호', 'ru': '루', 'mo': '모', 'ke': '커', 'ne': '네',
    'je': '제', 'yu': '유', 'se': '서', 'mi': '미', 'ju': '주', 'de': '데', 'oe': '외', 'wa': '와',
    'wi': '위', 'ri': '리', 'ye': '예', 'yi': '이', 'u': '우', 'eo': '어', 'heo': '허', 'du': '두',
    'leo': '러', 'lu': '루', 'so': '소', 'da': '다', 'lo': '로', 'nu': '누', 'o': '오', 'ga': '가'
}
dict_sorted = sorted(dict_map.items(), key=lambda x: len(x[0]), reverse=True)

# OCR 유틸 함수
def roman_to_korean(text):
    t = text.lower()
    for roman, kor in dict_sorted:
        t = t.replace(roman, kor)
    return t

def normalize(text):
    return re.sub(r'[^가-힣0-9]', '', text)

def is_valid_plate(text):
    return bool(re.fullmatch(r'\d{2,3}[가-힣]\d{4}', text))

def reorder_blocks(blocks):
    if len(blocks) == 2:
        if re.search(r'[가-힣]', blocks[1]) and not re.search(r'[가-힣]', blocks[0]):
            return [blocks[1], blocks[0]]
    return blocks

def insert_hangul_fixed(digits: str, hangul: str) -> str:
    if len(digits) >= 7:
        pos = -5
        return digits[:pos] + hangul + digits[pos+1:]
    return digits

def patch_hangul(t1, t2):
    d1 = ''.join(re.findall(r'\d', t1))
    d2 = ''.join(re.findall(r'\d', t2))
    h1 = re.findall(r'[가-힣]', t1)
    h2 = re.findall(r'[가-힣]', t2)
    if len(d1) in [7, 8] and len(h2) == 1:
        p = insert_hangul_fixed(d1, h2[0])
        if is_valid_plate(p): return p
    if len(d2) in [7, 8] and len(h1) == 1:
        p = insert_hangul_fixed(d2, h1[0])
        if is_valid_plate(p): return p
    return None

def get_filtered_ocr(reader, image, resize):
    resized = cv2.resize(image, resize)
    result = reader.readtext(resized)
    if not result:
        return '', 0.0
    blocks = reorder_blocks([t for (_, t, _) in result])
    merged = ''.join(blocks)
    norm = normalize(merged)
    conf = max([c for (_, _, c) in result])
    return norm, round(conf, 2)

def apply_plate_selection_logic(t1, c1, t2, c2, hangul_dict):
    if t1 == t2 and t1:
        return t1, '일치'
    if is_valid_plate(t1) and not is_valid_plate(t2):
        return t1, '정규식1'
    if is_valid_plate(t2) and not is_valid_plate(t1):
        return t2, '정규식2'
    if is_valid_plate(t1) and is_valid_plate(t2):
        h1, h2 = re.findall(r'[가-힣]', t1), re.findall(r'[가-힣]', t2)
        in1, in2 = h1 and h1[0] in hangul_dict, h2 and h2[0] in hangul_dict
        if in1 and not in2: return t1, '정규식+사전1'
        if in2 and not in1: return t2, '정규식+사전2'
        return (t1, '정규식+conf1') if c1 >= c2 else (t2, '정규식+conf2')
    patched = patch_hangul(t1, t2)
    if patched and is_valid_plate(patched):
        return patched, '패치'
    digits = re.sub(r'[^0-9]', '', t1)
    if len(digits) == 7:
        if digits[-5] == '4':
            p = insert_hangul_fixed(digits, '나')
            if is_valid_plate(p): return p, '보정(4-나)'
        if digits[-5] == '7':
            p = insert_hangul_fixed(digits, '가')
            if is_valid_plate(p): return p, '보정(7-가)'
    return (t1, 'conf1') if c1 >= c2 else (t2, 'conf2')

def get_plate_corners(image, fname=None, save_debug=False, debug_dir=None):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blur, 50, 150)

    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'gray.png'), gray)
        cv2.imwrite(os.path.join(debug_dir, 'blur.png'), blur)
        cv2.imwrite(os.path.join(debug_dir, 'canny.png'), edged)

    contours, _ = cv2.findContours(edged, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    contour_vis = image.copy()
    cv2.drawContours(contour_vis, contours, -1, (0, 255, 0), 1)
    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'contours.png'), contour_vis)

    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    for cnt in contours[:10]:
        area = cv2.contourArea(cnt)
        if area < 1000:
            continue
        approx = cv2.approxPolyDP(cnt, 0.02 * cv2.arcLength(cnt, True), True)
        if len(approx) == 4:
            pts = approx.reshape(4, 2)
            x, y, w, h = cv2.boundingRect(pts)
            if w / h < 2:
                continue
            return order_corners(pts)
    return None


def get_plate_corners_threshold(image, fname=None, save_debug=False, debug_dir=None):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 180, 255, cv2.THRESH_BINARY)

    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'blur_thresh.png'), blur)
        cv2.imwrite(os.path.join(debug_dir, 'thresh.png'), thresh)

    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contour_vis = image.copy()
    cv2.drawContours(contour_vis, contours, -1, (0, 0, 255), 1)
    if save_debug and debug_dir:
        cv2.imwrite(os.path.join(debug_dir, 'contours_thresh.png'), contour_vis)

    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < 1500:
            continue
        rect = cv2.minAreaRect(cnt)
        box = cv2.boxPoints(rect)
        ordered = order_corners(box.astype(np.float32))
        if save_debug and debug_dir:
            box_img = image.copy()
            cv2.polylines(box_img, [np.int32(ordered)], True, (255, 0, 0), 2)
            cv2.imwrite(os.path.join(debug_dir, 'minAreaRect_box.png'), box_img)
        return ordered
    return None



def order_corners(pts):
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]  # top-left
    rect[2] = pts[np.argmax(s)]  # bottom-right

    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]  # top-right
    rect[3] = pts[np.argmax(diff)]  # bottom-left
    return rect

def warp_perspective(image, corners, output_size=(200, 60)):
    dst = np.array([
        [0, 0],
        [output_size[0]-1, 0],
        [output_size[0]-1, output_size[1]-1],
        [0, output_size[1]-1]
    ], dtype="float32")
    M = cv2.getPerspectiveTransform(corners, dst)
    return cv2.warpPerspective(image, M, output_size)


# 모델 로딩
yolo_model = torch.hub.load('ultralytics/yolov5', 'custom',
                            path=os.path.join(BASE_DIR, 'custom_weights', 'best.pt'),
                            device='cpu', # 이 부분을 추가합니다.
                            verbose=False)

model_path = resource_path('custom_weights_easyOCR')
reader1 = easyocr.Reader(['ko'], gpu=False,
                         model_storage_directory=model_path,
                         user_network_directory=model_path,
                         recog_network='korean_g2', download_enabled=False)
reader2 = easyocr.Reader(['en'], gpu=False,
                         model_storage_directory=model_path,
                         user_network_directory=model_path,
                         recog_network='best_acc', download_enabled=False)

MIN_AREA = 1400
MIN_ASPECT_RATIO = 1.0
resize1 = (100, 32)
resize2 = (200, 60)
records = []

def detect_plate(image):
    detections = yolo_model(image).xyxy[0].cpu().numpy()
    filtered = []
    for *xyxy, conf, cls in detections.tolist():
        x1, y1, x2, y2 = map(int, xyxy)
        area = (x2 - x1) * (y2 - y1)
        ratio = (x2 - x1) / (y2 - y1 + 1e-5)
        if area > MIN_AREA and ratio > MIN_ASPECT_RATIO:
            filtered.append((area, (x1, y1, x2, y2)))
    if not filtered:
        return None
    _, (x1, y1, x2, y2) = max(filtered)
    return image[y1:y2, x1:x2], (x1, y1, x2, y2)


@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('images') or request.files.getlist('image')
    if not files:
        return jsonify({'error': '파일이 없습니다'}), 400

    for f in files:
        fname = secure_filename(f.filename)
        path = os.path.join(UPLOAD_FOLDER, fname)
        f.save(path)
        image = cv2.imread(path)
        result = {
            'image': f'/uploads/{fname}',
            'matched': False
        }

        # 디버그 폴더 생성
        debug_subdir = os.path.join(UPLOAD_FOLDER, 'debug', os.path.splitext(fname)[0])
        os.makedirs(debug_subdir, exist_ok=True)

        detected = detect_plate(image)
        if not detected:
            records.append(result)
            continue

        plate_img, (x1, y1, x2, y2) = detected

        # 디버그: YOLO 검출 영역 저장
        cv2.imwrite(os.path.join(debug_subdir, '1_detected_yolo.png'), plate_img)

        # 패딩 추가
        pad = 20
        x1 = max(x1 - pad, 0)
        y1 = max(y1 - pad, 0)
        x2 = min(x2 + pad, image.shape[1])
        y2 = min(y2 + pad, image.shape[0])
        plate_img = image[y1:y2, x1:x2]

        # 디버그: 패딩 후 이미지 저장
        cv2.imwrite(os.path.join(debug_subdir, '2_padded.png'), plate_img)

        # 디버그: 보정 전 이미지 저장
        cv2.imwrite(os.path.join(debug_subdir, '3_before_correction.png'), plate_img)

        # 보정 시도 1
        # corners 추출 시 디버그 인자 추가
        corners = get_plate_corners(plate_img, fname=os.path.splitext(fname)[0], save_debug=True, debug_dir=debug_subdir)
        if corners is None:
            corners = get_plate_corners_threshold(plate_img, fname=os.path.splitext(fname)[0], save_debug=True, debug_dir=debug_subdir)


        if corners is not None:
            # 디버그: 검출된 코너 시각화
            corner_img = plate_img.copy()
            cv2.polylines(corner_img, [np.int32(corners)], True, (0, 255, 0), 2)
            cv2.imwrite(os.path.join(debug_subdir, '4_detected_corners.png'), corner_img)

            # 보정 수행
            warped_img = warp_perspective(plate_img, corners)

            # 디버그: 보정된 결과 이미지 저장
            cv2.imwrite(os.path.join(debug_subdir, '5_corrected_warped.png'), warped_img)

            # 보정된 이미지로 교체
            plate_img = warped_img

        # OCR 리사이즈 후 입력 저장
        t1_input = cv2.resize(plate_img, resize1)
        t2_input = cv2.resize(plate_img, resize2)
        cv2.imwrite(os.path.join(debug_subdir, '6_ocr_model1_input.png'), t1_input)
        cv2.imwrite(os.path.join(debug_subdir, '7_ocr_model2_input.png'), t2_input)

        # OCR 수행
        t1, c1 = get_filtered_ocr(reader1, plate_img, resize1)
        t2, c2 = get_filtered_ocr(reader2, plate_img, resize2)
        selected, reason = apply_plate_selection_logic(t1, c1, t2, c2, set(dict_map.values()))
        matched = is_valid_plate(selected)

        if not matched:
            selected = '인식 실패'

        crop_name = f"crop_{fname}"
        vis_name = f"vis_{fname}"
        cv2.imwrite(os.path.join(CROP_FOLDER, crop_name), plate_img)
        PILImage.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB)).save(os.path.join(VISUAL_FOLDER, vis_name))

        # 개발자용 디버깅 이미지 저장: YOLO 박스 포함된 원본 이미지
        DEV_VISUAL_FOLDER = os.path.join(UPLOAD_FOLDER, 'dev_visual')
        os.makedirs(DEV_VISUAL_FOLDER, exist_ok=True)
        dev_vis_path = os.path.join(DEV_VISUAL_FOLDER, f'dev_{fname}')
        dev_img = image.copy()
        cv2.rectangle(dev_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        PILImage.fromarray(cv2.cvtColor(dev_img, cv2.COLOR_BGR2RGB)).save(dev_vis_path)

        result.update({
            'crop': f'/uploads/cropped/{crop_name}',
            'visual': f'/uploads/visual/{vis_name}',
            'text1': t1, 'conf1': c1,
            'text2': t2, 'conf2': c2,
            'plate': selected, 'reason': reason,
            'matched': matched
        })
        records.append(result)

    return jsonify({'status': 'ok'}), 200


@app.route('/download', methods=['GET'])
def download_excel():
    # 전체 records를 대상으로 함
    data = records.copy()

    if not data:
        return jsonify({'error': '데이터 없음'}), 400

    # matched가 False이면 결과를 "인식 실패"로 통일
    for r in data:
        if not r.get('matched'):
            r['text1'] = '인식 실패'
            r['text2'] = '인식 실패'
            r['plate'] = '인식 실패'

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df = pd.DataFrame(data)[['text1', 'text2', 'plate']]
        df.columns = ['모델1 결과', '모델2 결과', '선택된 결과']
        df.to_excel(writer, index=False, sheet_name='결과')
        ws = writer.sheets['결과']

        # 이미지 열 추가
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = '차량 이미지'

        # 16:9 비율 크기 설정
        TARGET_WIDTH = 150
        TARGET_HEIGHT = 267

        def px_to_col_width(px): return px * 0.14
        def px_to_row_height(px): return px * 0.75

        for idx, r in enumerate(data, start=2):
            try:
                # 원본 업로드된 이미지 경로로 변경
                img_path = os.path.join(BASE_DIR, r['image'].lstrip('/'))
                pil = PILImage.open(img_path)

                # (선택사항) 색상 모드 변환
                if pil.mode != 'RGB':
                    pil = pil.convert('RGB')

                # 원본 해상도 유지, 압축 없음
                img_bytes = io.BytesIO()
                pil.save(img_bytes, format='PNG')
                img_bytes.seek(0)

                xl_img = XLImage(img_bytes)
                xl_img.width = TARGET_WIDTH
                xl_img.height = TARGET_HEIGHT

                ws.add_image(xl_img, f'A{idx}')
                ws.row_dimensions[idx].height = px_to_row_height(TARGET_HEIGHT)

            except Exception as e:
                print(f"[이미지 삽입 실패] {r.get('image', '')} → {e}")



        ws.column_dimensions['A'].width = px_to_col_width(TARGET_WIDTH)

        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    buf.seek(0)
    today = datetime.now().strftime('%Y-%m-%d')
    fname = secure_filename(request.args.get('filename') or f"{today}_plates") + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/results')
def results():
    return jsonify(records), 200

@app.route('/update-plates', methods=['POST'])
def update_plates():
    updates = request.json  # 리스트 받기
    count = 0
    for u in updates:
        for r in records:
            if r['image'] == u['image']:
                r['plate'] = u['plate']
                r['matched'] = is_valid_plate(u['plate'])
                count += 1
    return jsonify({'updated': count}), 200



@app.route('/reset', methods=['POST'])
def reset():
    records.clear()
    return jsonify({'status': 'ok'})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)


import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import re
import io
import cv2
import torch
import easyocr
import numpy as np
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import warnings
import shutil

# 환경 설정
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "1"
warnings.filterwarnings("ignore", category=FutureWarning)
cv2.setNumThreads(0)
cv2.ocl.setUseOpenCL(False)

# 디렉토리 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
CROP_FOLDER = os.path.join(UPLOAD_FOLDER, 'cropped')
VISUAL_FOLDER = os.path.join(UPLOAD_FOLDER, 'visual')
for d in [UPLOAD_FOLDER, CROP_FOLDER, VISUAL_FOLDER]:
    os.makedirs(d, exist_ok=True)

app = Flask(__name__)
# Flask-CORS를 사용하여 CORS를 자동으로 처리합니다.
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 한글 매핑 딕셔너리
dict_map = {
    'gyeongnam': '경남', 'busan': '부산', 'jeo': '저', 'meo': '머', 'seo': '서', 'beo': '버',
    'geo': '거', 'neo': '너', 'deo': '더', 'bu': '부', 'do': '도', 'no': '노', 'go': '고', 'ro': '로',
    'bo': '보', 'jo': '조', 'gu': '구', 'na': '나', 'ma': '마', 'ba': '바', 'sa': '사', 'ah': '아',
    'ja': '자', 'cha': '차', 'ka': '카', 'ta': '타', 'pa': '파', 'ha': '하', 'la': '라', 'ra': '라',
    'me': '머', 'mu': '무', 'su': '수', 'ho': '호', 'ru': '루', 'mo': '모', 'ke': '커', 'ne': '네',
    'je': '제', 'yu': '유', 'se': '서', 'mi': '미', 'ju': '주', 'de': '데', 'oe': '외', 'wa': '와',
    'wi': '위', 'ri': '리', 'ye': '예', 'yi': '이', 'u': '우', 'eo': '어', 'heo': '허', 'du': '두',
    'leo': '러', 'lu': '루', 'so': '소', 'da': '다', 'lo': '로', 'nu': '누', 'o': '오', 'ga': '가'
}
dict_sorted = sorted(dict_map.items(), key=lambda x: len(x[0]), reverse=True)

# OCR 유틸 함수
def roman_to_korean(text):
    t = text.lower()
    for roman, kor in dict_sorted:
        t = t.replace(roman, kor)
    return t

def normalize(text):
    return re.sub(r'[^가-힣0-9]', '', text)

def is_valid_plate(text):
    return bool(re.fullmatch(r'\d{2,3}[가-힣]\d{4}', text))

def reorder_blocks(blocks):
    if len(blocks) == 2:
        if re.search(r'[가-힣]', blocks[1]) and not re.search(r'[가-힣]', blocks[0]):
            return [blocks[1], blocks[0]]
    return blocks

def insert_hangul_fixed(digits: str, hangul: str) -> str:
    if len(digits) >= 7:
        pos = -5
        return digits[:pos] + hangul + digits[pos+1:]
    return digits

def patch_hangul(t1, t2):
    d1 = ''.join(re.findall(r'\d', t1))
    d2 = ''.join(re.findall(r'\d', t2))
    h1 = re.findall(r'[가-힣]', t1)
    h2 = re.findall(r'[가-힣]', t2)
    if len(d1) in [7, 8] and len(h2) == 1:
        p = insert_hangul_fixed(d1, h2[0])
        if is_valid_plate(p): return p
    if len(d2) in [7, 8] and len(h1) == 1:
        p = insert_hangul_fixed(d2, h1[0])
        if is_valid_plate(p): return p
    return None

def get_filtered_ocr(reader, image, resize):
    resized = cv2.resize(image, resize)
    result = reader.readtext(resized)
    if not result:
        return '', 0.0
    blocks = reorder_blocks([t for (_, t, _) in result])
    merged = ''.join(blocks)
    norm = normalize(merged)
    conf = max([c for (_, _, c) in result])
    return norm, round(conf, 2)

def apply_plate_selection_logic(t1, c1, t2, c2, hangul_dict):
    if t1 == t2 and t1:
        return t1, '일치'
    if is_valid_plate(t1) and not is_valid_plate(t2):
        return t1, '정규식1'
    if is_valid_plate(t2) and not is_valid_plate(t1):
        return t2, '정규식2'
    if is_valid_plate(t1) and is_valid_plate(t2):
        h1, h2 = re.findall(r'[가-힣]', t1), re.findall(r'[가-힣]', t2)
        in1, in2 = h1 and h1[0] in hangul_dict, h2 and h2[0] in hangul_dict
        if in1 and not in2: return t1, '정규식+사전1'
        if in2 and not in1: return t2, '정규식+사전2'
        return (t1, '정규식+conf1') if c1 >= c2 else (t2, '정규식+conf2')
    patched = patch_hangul(t1, t2)
    if patched and is_valid_plate(patched):
        return patched, '패치'
    digits = re.sub(r'[^0-9]', '', t1)
    if len(digits) == 7:
        if digits[-5] == '4':
            p = insert_hangul_fixed(digits, '나')
            if is_valid_plate(p): return p, '보정(4-나)'
        if digits[-5] == '7':
            p = insert_hangul_fixed(digits, '가')
            if is_valid_plate(p): return p, '보정(7-가)'
    return (t1, 'conf1') if c1 >= c2 else (t2, 'conf2')

# 모델 로딩
yolo_model = torch.hub.load('ultralytics/yolov5', 'custom',
                             path=os.path.join(BASE_DIR, 'custom_weights', 'best.pt'),
                             force_reload=False, verbose=False)
reader1 = easyocr.Reader(['ko'], gpu=False,
                         model_storage_directory='custom_weights_easyOCR',
                         user_network_directory='custom_weights_easyOCR',
                         recog_network='korean_g2', download_enabled=False)
reader2 = easyocr.Reader(['en'], gpu=False,
                         model_storage_directory='custom_weights_easyOCR',
                         user_network_directory='custom_weights_easyOCR',
                         recog_network='english_g2', download_enabled=False)

MIN_AREA = 1400
MIN_ASPECT_RATIO = 1.0
resize1 = (100, 32)
resize2 = (200, 60)
records = []

def detect_plate(image):
    detections = yolo_model(image).xyxy[0].cpu().numpy()
    filtered = []
    for *xyxy, conf, cls in detections.tolist():
        x1, y1, x2, y2 = map(int, xyxy)
        area = (x2 - x1) * (y2 - y1)
        ratio = (x2 - x1) / (y2 - y1 + 1e-5)
        if area > MIN_AREA and ratio > MIN_ASPECT_RATIO:
            filtered.append((area, (x1, y1, x2, y2)))
    if not filtered:
        return None
    _, (x1, y1, x2, y2) = max(filtered)
    return image[y1:y2, x1:x2], (x1, y1, x2, y2)

@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('images') or request.files.getlist('image')
    if not files:
        return jsonify({'error': '파일이 없습니다'}), 400

    for f in files:
        fname = secure_filename(f.filename)
        path = os.path.join(UPLOAD_FOLDER, fname)
        f.save(path)
        image = cv2.imread(path)
        result = {
            'image': f'/uploads/{fname}',
            'matched': False
        }

        detected = detect_plate(image)
        if not detected:
            records.append(result)
            continue

        plate_img, (x1, y1, x2, y2) = detected
        t1, c1 = get_filtered_ocr(reader1, plate_img, resize1)
        t2, c2 = get_filtered_ocr(reader2, plate_img, resize2)
        selected, reason = apply_plate_selection_logic(t1, c1, t2, c2, set(dict_map.values()))
        matched = is_valid_plate(selected)

        # 인식 실패 처리
        if not matched:
            selected = '인식 실패'
            # 인식 실패 시 정답률은 0.0으로 처리
            c1 = 0.0
            c2 = 0.0


        crop_name = f"crop_{fname}"
        vis_name = f"vis_{fname}"
        cv2.imwrite(os.path.join(CROP_FOLDER, crop_name), plate_img)
        vis = image.copy()
        cv2.rectangle(vis, (x1, y1), (x2, y2), (0,255,0), 2)
        PILImage.fromarray(cv2.cvtColor(vis, cv2.COLOR_BGR2RGB)).save(os.path.join(VISUAL_FOLDER, vis_name))

        result.update({
            'crop': f'/uploads/cropped/{crop_name}',
            'visual': f'/uploads/visual/{vis_name}',
            'text1': t1, 'conf1': c1,
            'text2': t2, 'conf2': c2,
            'plate': selected, 'reason': reason,
            'matched': matched
        })
        records.append(result)

    return jsonify({'status': 'ok'}), 200

@app.route('/download', methods=['GET'])
def download_excel():
    data = records.copy()

    if not data:
        return jsonify({'error': '데이터 없음'}), 400

    for r in data:
        if not r.get('matched'):
            r['text1'] = '인식 실패'
            r['text2'] = '인식 실패'
            r['plate'] = '인식 실패'

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df = pd.DataFrame(data)[['text1', 'text2', 'plate']]
        df.columns = ['모델1 결과', '모델2 결과', '선택된 결과']
        df.to_excel(writer, index=False, sheet_name='결과')
        ws = writer.sheets['결과']

        # 첫 번째 열에 이미지 삽입을 위해 열을 추가합니다.
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = '차량 이미지'

        TARGET_WIDTH = 150
        TARGET_HEIGHT = 267

        def px_to_col_width(px): return px * 0.14
        def px_to_row_height(px): return px * 0.75

        for idx, r in enumerate(data, start=2):
            try:
                # 'visual' 키가 있는지 확인하여 에러를 방지합니다.
                if 'visual' in r and r['visual']:
                    img_path = os.path.join(BASE_DIR, r['visual'].lstrip('/'))
                    
                    if os.path.exists(img_path):
                        pil = PILImage.open(img_path)

                        pil = pil.resize((TARGET_WIDTH, TARGET_HEIGHT), PILImage.LANCZOS)
                        img_bytes = io.BytesIO()
                        pil.save(img_bytes, format='PNG')
                        img_bytes.seek(0)

                        xl_img = XLImage(img_bytes)
                        xl_img.width = TARGET_WIDTH
                        xl_img.height = TARGET_HEIGHT

                        ws.add_image(xl_img, f'A{idx}')
                        ws.row_dimensions[idx].height = px_to_row_height(TARGET_HEIGHT)
                    else:
                        print(f"이미지 파일이 존재하지 않습니다: {img_path}")
                else:
                    print(f"이미지 경로가 데이터에 없습니다: {r}")
            except Exception as e:
                print(f"[이미지 삽입 실패] {r.get('visual', '경로 없음')} → {e}")

        ws.column_dimensions['A'].width = px_to_col_width(TARGET_WIDTH)

        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    buf.seek(0)
    today = datetime.now().strftime('%Y-%m-%d')
    fname = secure_filename(request.args.get('filename') or f"{today}_plates") + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download-json', methods=['GET'])
def download_json():
    # raw 데이터를 포함한 JSON 리스트를 생성합니다.
    json_data = []
    for i, r in enumerate(records):
        accuracy_rate = r['conf1'] if r['plate'] == r['text1'] else r['conf2']
        item = {
            '연번': i + 1,
            '날짜': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '사진의이름 및 주소': r.get('image', '정보 없음'),
            '인식성공 or 실패': '성공' if r.get('matched', False) else '실패',
            '정답률': f"{accuracy_rate * 100:.2f}%"
        }
        json_data.append(item)
    
    # JSON 파일로 다운로드하도록 응답을 구성합니다.
    json_buf = io.BytesIO()
    json_buf.write(jsonify(json_data).data)
    json_buf.seek(0)
    
    today = datetime.now().strftime('%Y-%m-%d')
    fname = secure_filename(request.args.get('filename') or f"{today}_plates") + '.json'
    
    return send_file(json_buf, as_attachment=True, download_name=fname, mimetype='application/json')

@app.route('/results')
def results():
    return jsonify(records), 200

@app.route('/update-plates', methods=['POST'])
def update_plates():
    updates = request.json
    count = 0
    for u in updates:
        for r in records:
            if r['image'] == u['image']:
                r['plate'] = u['plate']
                r['matched'] = is_valid_plate(u['plate'])
                count += 1
    return jsonify({'updated': count}), 200

@app.route('/reset', methods=['POST'])
def reset():
    records.clear()
    shutil.rmtree(UPLOAD_FOLDER, ignore_errors=True)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(CROP_FOLDER, exist_ok=True)
    os.makedirs(VISUAL_FOLDER, exist_ok=True)
    return jsonify({'status': 'ok'}), 200

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/uploads/cropped/<path:filename>')
def cropped_file(filename):
    return send_from_directory(CROP_FOLDER, filename)

@app.route('/uploads/visual/<path:filename>')
def visual_file(filename):
    return send_from_directory(VISUAL_FOLDER, filename)


@app.route('/image/<path:filename>', methods=['DELETE'])
def delete_image(filename):
    global records
    # records에서 해당 파일의 모든 경로를 찾습니다.
    # filename 인자에서 실제 파일 이름만 추출합니다.
    base_filename = os.path.basename(filename)
    item = next((r for r in records if r['image'] == f'/uploads/{base_filename}'), None)
    
    if item:
        # 원본, 잘라낸 이미지, 시각화 이미지의 경로를 만듭니다.
        original_path = os.path.join(UPLOAD_FOLDER, base_filename)
        crop_path = os.path.join(CROP_FOLDER, f"crop_{base_filename}")
        vis_path = os.path.join(VISUAL_FOLDER, f"vis_{base_filename}")
        
        # 파일들을 삭제합니다. (파일이 없을 경우 오류 방지를 위해 os.path.exists()를 사용)
        try:
            if os.path.exists(original_path):
                os.remove(original_path)
            if os.path.exists(crop_path):
                os.remove(crop_path)
            if os.path.exists(vis_path):
                os.remove(vis_path)
        except OSError as e:
            print(f"파일 삭제 중 오류 발생: {e}")
            return jsonify({'error': 'Failed to delete files'}), 500

        # records 리스트에서 해당 항목을 제거합니다.
        records.remove(item)
        
        return jsonify({'message': f'File {base_filename} and associated data deleted successfully'}), 200
    else:
        return jsonify({'error': 'File not found in records'}), 404

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
